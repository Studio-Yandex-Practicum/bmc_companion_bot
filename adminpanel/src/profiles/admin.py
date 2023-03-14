from datetime import datetime

import openpyxl
from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.http import HttpResponse
from questioning.models import TestCompleted
from schedule.models import MeetingFeedback

from . import models

EXPORT_FIELDS = ("id", "first_name", "last_name", "phone", "telegram_login", "age")
TEST_EXPORT_FIELDS = ("Название теста", "Баллы")
FEEDBACK_EXPORT_FIELDS = (
    "Дата",
    "Время",
    "Фамилия психолога",
    "Имя психолога",
    "Телеграм психолога",
    "Отзыв",
    "Оценка",
)


@admin.action(description="Загрузить как файл *.xlsx")
def export_table_to_excel(modeladmin, request, queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    qs_fields = queryset.model._meta.fields
    user_fields = [f.name for f in qs_fields if f.name in EXPORT_FIELDS]
    user_fields_ru = [f.verbose_name for f in qs_fields if f.name in EXPORT_FIELDS]
    test_results = TestCompleted.objects.all()
    feedbacks = MeetingFeedback.objects.all()
    row = 1
    for user in queryset.values_list(*user_fields):
        user_test_results = test_results.filter(profile=int(user[0]))
        user_feedbacks = feedbacks.filter(user=int(user[0]))
        for col_num, field_name in enumerate(user_fields_ru):
            if col_num > 0:
                ws.cell(row=row, column=col_num).value = str(field_name).capitalize()
        row += 1
        for col_num, field_value in enumerate(user):
            if col_num > 0:
                ws.cell(row=row, column=col_num).value = str(field_value)
        row += 2
        if user_test_results:
            for col_num, field_name in enumerate(TEST_EXPORT_FIELDS):
                ws.cell(row=row, column=col_num + 1).value = str(field_name)
            row += 1
            for row_num, test in enumerate(user_test_results):
                ws.cell(row=row + row_num, column=1).value = str(test.test.name)
                ws.cell(row=row + row_num, column=2).value = str(test.value)
            row += 2
        if user_feedbacks:
            for col_num, field_name in enumerate(FEEDBACK_EXPORT_FIELDS):
                ws.cell(row=row, column=col_num + 1).value = str(field_name)
            row += 1
            for row_num, feedback in enumerate(user_feedbacks):
                ws.cell(row=row + row_num, column=1).value = str(
                    feedback.meeting.date_start.strftime("%d.%m.%Y")
                )
                ws.cell(row=row + row_num, column=2).value = str(
                    feedback.meeting.date_start.strftime("%H:%M")
                )
                ws.cell(row=row + row_num, column=3).value = str(
                    feedback.meeting.psychologist.last_name
                )
                ws.cell(row=row + row_num, column=4).value = str(
                    feedback.meeting.psychologist.first_name
                )
                ws.cell(row=row + row_num, column=5).value = str(
                    feedback.meeting.psychologist.telegram_login
                )
                ws.cell(row=row + row_num, column=6).value = str(feedback.text)
                ws.cell(row=row + row_num, column=7).value = str(feedback.score)
            row += 2
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=profile-{datetime.now()}.xlsx"
    wb.save(response)
    return response


@admin.register(models.Profile)
class ProfileAdmin(UserAdmin):
    list_display = ("id", "first_name", "last_name", "telegram_login", "phone")
    search_fields = (
        "username",
        "email",
        "first_name",
        "last_name",
        "telegram_login",
        "phone",
        "telegram_id",
        "chat_id",
    )
    actions = [export_table_to_excel]
    list_filter = ("is_staff", "is_active")
    fieldsets = (
        (
            "Учетные данные",
            {"fields": ("username", "password", "email")},
        ),
        (
            "Основное",
            {
                "fields": (
                    "first_name",
                    "last_name",
                    "middle_name",
                    "birthday",
                    "phone",
                    "age",
                    "uce_score",
                )
            },
        ),
        (
            "Telegram",
            {"fields": ("telegram_login", "telegram_id", "chat_id")},
        ),
        ("Персонал", {"fields": ("is_staff", "is_active", "is_superuser")}),
        ("Полномочия", {"fields": ("groups", "user_permissions")}),
        (
            "Прочее",
            {
                "fields": (
                    "date_joined",
                    "last_login",
                )
            },
        ),
    )
